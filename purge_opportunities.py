#!/usr/bin/env python3
"""Purge the opportunities tracker before a fresh discovery run.

The opportunities xlsx has two sheets: `Opportunities` (the active backlog) and
`Rejected opportunities` (an append-only memory of roles you reviewed and said
no to, so they never resurface). This script sweeps the active sheet:

  - Rows where `Worth Applying` = N or `Valid?` = N  → moved to the rejected
    sheet with today's date and a one-line reason.
  - Rows where `Applied` = Y/Yes                     → deleted (already applied).
  - Surviving rows are renumbered so `S.No` stays sequential.

Run:
  python3 purge_opportunities.py
  python3 purge_opportunities.py --opps "TPM opportunities.xlsx"
"""
from __future__ import annotations

import argparse
import datetime as dt
import os

from openpyxl import load_workbook


def is_yes(v) -> bool:
    return v is not None and str(v).strip().lower() in {"y", "yes"}


def is_no(v) -> bool:
    return v is not None and str(v).strip().lower() in {"n", "no"}


def sync_rejected_header(opp_header: list, rej_header: list) -> list:
    """Return the rejected sheet's header, widened to cover every named column in
    `opp_header` plus `Rejected on` / `Reason`.

    The two sheets drift as the active sheet gains columns over time while the
    rejected header stays frozen. Widening by name (never by position) means a
    moved row never loses a column or lands its reason in the wrong cell.
    """
    out = [str(h).strip() if h is not None else "" for h in rej_header]
    while out and not out[-1]:  # drop trailing blank padding before appending
        out.pop()
    for name in list(opp_header) + ["Rejected on", "Reason"]:
        if name is None or not str(name).strip():
            continue
        name = str(name).strip()
        if not any(h.lower() == name.lower() for h in out):
            out.append(name)
    return out


def build_rejected_row(opp_header: list, vals: list, rej_header: list,
                       today: str, reason: str) -> list:
    """Map an Opportunities row onto the rejected sheet's columns BY NAME.

    Positional concatenation (`vals + [today, reason]`) silently breaks once the
    two headers differ in width — the reason lands in the wrong cell. Mapping by
    name keeps `Rejected on` / `Reason` aligned no matter how the sheets drift.
    """
    lookup = {}
    for i, name in enumerate(opp_header):
        if name is not None and str(name).strip():
            lookup[str(name).strip().lower()] = vals[i] if i < len(vals) else None
    lookup["rejected on"] = today
    lookup["reason"] = reason
    out = [None] * len(rej_header)
    for i, name in enumerate(rej_header):
        key = str(name).strip().lower() if name else ""
        if key in lookup:
            out[i] = lookup[key]
    return out


def purge(path: str) -> dict:
    wb = load_workbook(path)
    if "Opportunities" not in wb.sheetnames:
        raise SystemExit(f"missing sheet 'Opportunities' in {path}")
    opp = wb["Opportunities"]
    if "Rejected opportunities" not in wb.sheetnames:
        wb.create_sheet("Rejected opportunities")
    rej = wb["Rejected opportunities"]

    opp_header = [c.value for c in opp[1]]

    # Build or widen the rejected header to cover the active sheet's columns.
    if rej.max_row == 0 or all(c.value is None for c in rej[1]):
        rej_header = sync_rejected_header(opp_header, [])
    else:
        rej_header = sync_rejected_header(opp_header, [c.value for c in rej[1]])
    for i, name in enumerate(rej_header, start=1):
        rej.cell(row=1, column=i, value=name)

    def col(*names: str) -> int:
        lowered = [str(h).strip().lower() if h is not None else "" for h in opp_header]
        for name in names:
            if name.lower() in lowered:
                return lowered.index(name.lower()) + 1
        return -1

    worth_col = col("Worth Applying")
    valid_col = col("Valid?")
    applied_col = col("Applied")
    sno_col = col("S.No", "S.No.")
    company_col = col("Company name", "Company")
    title_col = col("Title")
    today = dt.date.today().isoformat()

    keep_rows: list[list] = []
    rejected_count = applied_dropped = 0

    for row in opp.iter_rows(min_row=2, values_only=False):
        vals = [c.value for c in row]
        cell = lambda ci: vals[ci - 1] if ci > 0 and ci - 1 < len(vals) else None
        company, title = cell(company_col) or "", cell(title_col) or ""

        if is_yes(cell(applied_col)):
            applied_dropped += 1
            continue

        if is_no(cell(worth_col)) or is_no(cell(valid_col)):
            reasons = []
            if is_no(cell(worth_col)):
                reasons.append("Worth Applying = N")
            if is_no(cell(valid_col)):
                reasons.append("Valid = N")
            rej.append(build_rejected_row(opp_header, vals, rej_header, today, "; ".join(reasons)))
            rejected_count += 1
            continue

        keep_rows.append(list(vals))

    if opp.max_row > 1:
        opp.delete_rows(2, opp.max_row - 1)
    for i, vals in enumerate(keep_rows, start=1):
        if sno_col > 0:
            vals[sno_col - 1] = i
        opp.append(vals)

    wb.save(path)
    return {"kept": len(keep_rows), "rejected_moved": rejected_count,
            "applied_dropped": applied_dropped}


def main() -> None:
    here = os.path.dirname(os.path.abspath(__file__))
    p = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    p.add_argument("--opps", default=os.path.join(here, "TPM opportunities.xlsx"),
                   help="Path to the opportunities xlsx (default: sibling file).")
    args = p.parse_args()
    result = purge(os.path.abspath(args.opps))
    print(f"kept={result['kept']}, rejected_moved={result['rejected_moved']}, "
          f"applied_dropped={result['applied_dropped']}")


if __name__ == "__main__":
    main()
