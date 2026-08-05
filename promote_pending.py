#!/usr/bin/env python3
"""Promote staged applications from Pending-applications/ once you've applied.

The folder's location is the state: a role you're preparing sits in
`Pending-applications/<Company>/` with its JD. The moment you drop your tailored
resume (any file with "resume" in the name) into that folder, the role is
applied — so this script graduates it:

  1. Moves `Pending-applications/<Company>/` up to `<root>/<Company>/`
     (the applied set). If a folder for that company already exists, its files
     are merged in.
  2. Writes an application-tracker row per JD in the folder — `Date applied` =
     today, `Status` = "Applied - awaiting response" — carrying `Title`/`Link`
     from the matching opportunities row when one exists. An existing row for the
     same company+role is updated in place rather than duplicated.
  3. Flags the matching opportunities row `Applied = Y` so the next backlog purge
     drops it.

Deterministic file + spreadsheet ops — no network, no connector. A folder with
no resume file is left untouched.

Run:
  python3 promote_pending.py
  python3 promote_pending.py --root Job-applications-TPM
"""
from __future__ import annotations

import argparse
import datetime as dt
import os
import shutil
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from normalize import normalize_company, normalize_title  # noqa: E402
from read_jds import RESUME_RX, COVER_RX, normalize_filename_title  # noqa: E402

PENDING_DIRNAME = "Pending-applications"
APPLIED_STATUS = "Applied - awaiting response"


def _header_index(header, *names):
    lowered = [str(h).strip().lower() if h is not None else "" for h in header]
    for name in names:
        if name.lower() in lowered:
            return lowered.index(name.lower())
    return None


def _jd_files(folder: str) -> list[str]:
    """JD .docx files in a folder — excludes resume and cover-letter files."""
    out = []
    for f in sorted(os.listdir(folder)):
        if not f.lower().endswith(".docx") or f.startswith((".", "~$")):
            continue
        if RESUME_RX.search(f) or COVER_RX.search(f):
            continue
        out.append(f)
    return out


def _has_resume(folder: str) -> bool:
    return any(
        RESUME_RX.search(f) and not f.startswith((".", "~$"))
        for f in os.listdir(folder)
    )


def _load_opps_index(opps_path: str):
    """Return (workbook, sheet, header, {(norm_company, norm_title): row_idx}, link_by_key)
    for the opportunities Opportunities sheet, or (None, ...) if unavailable."""
    if not os.path.isfile(opps_path):
        return None, None, None, {}, {}
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("promote_pending: openpyxl not installed; skipping opps flag", file=sys.stderr)
        return None, None, None, {}, {}
    wb = load_workbook(opps_path)
    ws = wb["Opportunities"] if "Opportunities" in wb.sheetnames else wb.active
    header = [c.value for c in ws[1]]
    ci = _header_index(header, "Company name", "Company")
    ti = _header_index(header, "Title")
    li = _header_index(header, "Link")
    index, link_by_key = {}, {}
    if ci is not None and ti is not None:
        for r in range(2, ws.max_row + 1):
            comp = ws.cell(row=r, column=ci + 1).value
            title = ws.cell(row=r, column=ti + 1).value
            if not comp:
                continue
            key = (normalize_company(str(comp)), normalize_title(str(title or "")))
            index[key] = r
            if li is not None:
                link_by_key[key] = ws.cell(row=r, column=li + 1).value
    return wb, ws, header, index, link_by_key


def _append_or_update_tracker(ws, header, company, title, link, today):
    ci = _header_index(header, "Company name", "Company")
    ti = _header_index(header, "Title")
    li = _header_index(header, "Link")
    di = _header_index(header, "Date applied")
    si = _header_index(header, "Status")
    sno_i = _header_index(header, "S.No", "S.No.")
    key = (normalize_company(company), normalize_title(title))

    # Update in place if this company+role already has a row.
    for r in range(2, ws.max_row + 1):
        rc = ws.cell(row=r, column=ci + 1).value if ci is not None else None
        rt = ws.cell(row=r, column=ti + 1).value if ti is not None else None
        if rc and (normalize_company(str(rc)), normalize_title(str(rt or ""))) == key:
            if di is not None:
                ws.cell(row=r, column=di + 1, value=today)
            if si is not None:
                ws.cell(row=r, column=si + 1, value=APPLIED_STATUS)
            return "updated"

    row = [None] * len(header)
    def put(idx, val):
        if idx is not None and val is not None:
            row[idx] = val
    put(ci, company)
    put(ti, title)
    put(li, link)
    put(di, today)
    put(si, APPLIED_STATUS)
    if sno_i is not None:
        max_sno = 0
        for r in range(2, ws.max_row + 1):
            v = ws.cell(row=r, column=sno_i + 1).value
            if isinstance(v, int):
                max_sno = max(max_sno, v)
        row[sno_i] = max_sno + 1
    ws.append(row)
    return "appended"


def _move_folder(src: str, dst: str):
    if not os.path.exists(dst):
        shutil.move(src, dst)
        return
    for f in os.listdir(src):  # dst exists → merge files in, skip name clashes
        target = os.path.join(dst, f)
        if not os.path.exists(target):
            shutil.move(os.path.join(src, f), target)
    if not os.listdir(src):
        os.rmdir(src)


def promote(root: str, tracker_path: str, opps_path: str, today: str | None = None) -> dict:
    today = today or dt.date.today().isoformat()
    pending_root = os.path.join(root, PENDING_DIRNAME)
    result = {"promoted": [], "skipped_no_resume": [], "tracker_rows": 0}
    if not os.path.isdir(pending_root):
        return result

    try:
        from openpyxl import load_workbook
    except ImportError:
        print("promote_pending: openpyxl not installed; cannot update tracker", file=sys.stderr)
        load_workbook = None

    tracker_wb = tracker_ws = tracker_header = None
    if load_workbook and os.path.isfile(tracker_path):
        tracker_wb = load_workbook(tracker_path)
        tracker_ws = tracker_wb.active
        tracker_header = [c.value for c in tracker_ws[1]]

    opps_wb, opps_ws, _, opps_index, opps_link = _load_opps_index(opps_path)

    for company in sorted(os.listdir(pending_root)):
        src = os.path.join(pending_root, company)
        if not os.path.isdir(src) or company.startswith("."):
            continue
        if not _has_resume(src):
            result["skipped_no_resume"].append(company)
            continue

        jds = _jd_files(src)
        # Resolve tracker rows BEFORE the move (read titles from the folder).
        for jd in jds or [None]:
            title = os.path.splitext(jd)[0] if jd else company
            key = (normalize_company(company),
                   normalize_title(normalize_filename_title(jd)) if jd else normalize_company(company))
            link = opps_link.get(key)
            if tracker_ws is not None:
                _append_or_update_tracker(tracker_ws, tracker_header, company, title, link, today)
                result["tracker_rows"] += 1
            if opps_ws is not None and key in opps_index:
                ai = _header_index([c.value for c in opps_ws[1]], "Applied")
                if ai is not None:
                    opps_ws.cell(row=opps_index[key], column=ai + 1, value="Y")

        _move_folder(src, os.path.join(root, company))
        result["promoted"].append(company)

    if tracker_wb is not None:
        tracker_wb.save(tracker_path)
    if opps_wb is not None:
        opps_wb.save(opps_path)
    return result


def main():
    here = os.path.dirname(os.path.abspath(__file__))
    p = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    p.add_argument("--root", default=here,
                   help="Directory holding Pending-applications/ and per-company folders "
                        "(default: this script's dir).")
    p.add_argument("--tracker", default=os.path.join(here, "Application tracker.xlsx"))
    p.add_argument("--opps", default=os.path.join(here, "TPM opportunities.xlsx"))
    args = p.parse_args()
    r = promote(os.path.abspath(args.root), os.path.abspath(args.tracker), os.path.abspath(args.opps))
    print(f"promoted={r['promoted']}, skipped_no_resume={r['skipped_no_resume']}, "
          f"tracker_rows_written={r['tracker_rows']}")


if __name__ == "__main__":
    main()
