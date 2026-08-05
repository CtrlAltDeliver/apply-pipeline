"""Tests for promote_pending.py — graduate applied folders + write the tracker.

Runnable two ways:
    pytest
    python3 tests/test_promote_pending.py
"""
import os
import sys
import tempfile

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from openpyxl import Workbook, load_workbook  # noqa: E402
from promote_pending import promote  # noqa: E402

TRACKER_HEADER = ["Company", "Title", "Link", "Date applied", "Status"]
OPP_HEADER = ["S.No", "Title", "Company name", "Location", "Link", "Fitment",
              "Worth Applying", "Valid?", "Applied", "Who can refer"]


def _tracker(path, rows=()):
    wb = Workbook()
    ws = wb.active
    ws.title = "Applications"
    ws.append(TRACKER_HEADER)
    for r in rows:
        ws.append(r)
    wb.save(path)


def _opps(path, rows=()):
    wb = Workbook()
    ws = wb.active
    ws.title = "Opportunities"
    ws.append(OPP_HEADER)
    for r in rows:
        ws.append(r)
    wb.create_sheet("Rejected opportunities")
    wb.save(path)


def _pending(root, company, files):
    d = os.path.join(root, "Pending-applications", company)
    os.makedirs(d)
    for f in files:
        open(os.path.join(d, f), "w").close()


def test_folder_with_resume_is_promoted_and_tracked():
    with tempfile.TemporaryDirectory() as root:
        tracker = os.path.join(root, "Application tracker.xlsx")
        opps = os.path.join(root, "TPM opportunities.xlsx")
        _tracker(tracker)
        _opps(opps, [[1, "Senior TPM", "Acme", "Remote", "https://acme/jobs/9",
                      "4", "Y", "Y", "", ""]])
        _pending(root, "Acme", ["Senior TPM.docx", "Jane Doe Resume.docx"])

        r = promote(root, tracker, opps, today="2026-02-01")
        assert r["promoted"] == ["Acme"]
        assert r["tracker_rows"] == 1

        # Folder moved out of Pending, up to the applied set.
        assert not os.path.exists(os.path.join(root, "Pending-applications", "Acme"))
        assert os.path.isdir(os.path.join(root, "Acme"))

        # Tracker row written with today's date, status, and the opps Link.
        tw = load_workbook(tracker).active
        row = [c.value for c in list(tw.iter_rows())[1]]
        assert row[0] == "Acme" and row[1] == "Senior TPM"
        assert row[2] == "https://acme/jobs/9"
        assert row[3] == "2026-02-01"
        assert row[4] == "Applied - awaiting response"

        # Opportunities row flagged Applied = Y.
        ow = load_workbook(opps)["Opportunities"]
        applied_col = OPP_HEADER.index("Applied")
        assert list(ow.iter_rows(values_only=True))[1][applied_col] == "Y"


def test_folder_without_resume_is_skipped():
    with tempfile.TemporaryDirectory() as root:
        tracker = os.path.join(root, "Application tracker.xlsx")
        opps = os.path.join(root, "TPM opportunities.xlsx")
        _tracker(tracker)
        _opps(opps)
        _pending(root, "BetaCo", ["JD.docx"])  # no resume

        r = promote(root, tracker, opps, today="2026-02-01")
        assert r["promoted"] == []
        assert r["skipped_no_resume"] == ["BetaCo"]
        assert os.path.isdir(os.path.join(root, "Pending-applications", "BetaCo"))
        assert len(list(load_workbook(tracker).active.iter_rows())) == 1  # header only


def test_existing_tracker_row_is_updated_not_duplicated():
    with tempfile.TemporaryDirectory() as root:
        tracker = os.path.join(root, "Application tracker.xlsx")
        opps = os.path.join(root, "TPM opportunities.xlsx")
        # A pre-existing row for the same company+role, stale status.
        _tracker(tracker, [["Acme", "Sr. Technical Program Manager", "", "", "Interested"]])
        _opps(opps)
        _pending(root, "Acme", ["Senior Technical Program Manager.docx", "resume.pdf"])

        promote(root, tracker, opps, today="2026-03-03")
        rows = list(load_workbook(tracker).active.iter_rows(values_only=True))
        assert len(rows) == 2  # header + the one (updated, not duplicated) row
        assert rows[1][3] == "2026-03-03"
        assert rows[1][4] == "Applied - awaiting response"


def test_merges_into_existing_company_folder():
    with tempfile.TemporaryDirectory() as root:
        tracker = os.path.join(root, "Application tracker.xlsx")
        opps = os.path.join(root, "TPM opportunities.xlsx")
        _tracker(tracker)
        _opps(opps)
        # Company already has an applied folder from a prior role.
        os.makedirs(os.path.join(root, "Acme"))
        open(os.path.join(root, "Acme", "Old Role.docx"), "w").close()
        _pending(root, "Acme", ["New Role.docx", "resume.docx"])

        promote(root, tracker, opps, today="2026-02-01")
        applied = set(os.listdir(os.path.join(root, "Acme")))
        assert {"Old Role.docx", "New Role.docx", "resume.docx"} <= applied
        assert not os.path.exists(os.path.join(root, "Pending-applications", "Acme"))


if __name__ == "__main__":
    fns = [v for k, v in sorted(globals().items()) if k.startswith("test_") and callable(v)]
    failures = []
    for fn in fns:
        try:
            fn()
        except AssertionError as e:
            failures.append(f"{fn.__name__}: {e}")
    if failures:
        print("FAILED:")
        for f in failures:
            print("  -", f)
        sys.exit(1)
    print(f"all {len(fns)} promote_pending tests passed")
