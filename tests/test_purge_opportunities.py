"""Tests for purge_opportunities.py — the opportunities-tracker sweep.

Builds a small workbook in a temp dir, purges it, and asserts the rows landed on
the right sheet with columns aligned by name.

Runnable two ways:
    pytest
    python3 tests/test_purge_opportunities.py
"""
import os
import sys
import tempfile

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from openpyxl import Workbook, load_workbook  # noqa: E402
from purge_opportunities import purge, sync_rejected_header, build_rejected_row  # noqa: E402

OPP_HEADER = ["S.No", "Title", "Company name", "Location", "Link",
              "Fitment", "Worth Applying", "Valid?", "Applied", "Who can refer"]


def _make_workbook(path, rows):
    wb = Workbook()
    opp = wb.active
    opp.title = "Opportunities"
    opp.append(OPP_HEADER)
    for r in rows:
        opp.append(r)
    wb.create_sheet("Rejected opportunities")
    wb.save(path)


def _row(sno, title, company, worth="", valid="", applied=""):
    return [sno, title, company, "Remote", f"https://x/{sno}", "4",
            worth, valid, applied, ""]


def test_purge_moves_and_deletes_correctly():
    with tempfile.TemporaryDirectory() as d:
        path = os.path.join(d, "opps.xlsx")
        _make_workbook(path, [
            _row(1, "Keeper TPM", "GoodCo", worth="Y", valid="Y"),
            _row(2, "Not Worth", "MehCo", worth="N", valid="Y"),
            _row(3, "Invalid", "BadCo", worth="Y", valid="N"),
            _row(4, "Already Applied", "DoneCo", worth="Y", valid="Y", applied="Y"),
            _row(5, "Undecided", "MaybeCo"),  # blank worth/valid -> kept
        ])
        result = purge(path)
        assert result == {"kept": 2, "rejected_moved": 2, "applied_dropped": 1}

        wb = load_workbook(path)
        opp_titles = [r[1] for r in wb["Opportunities"].iter_rows(min_row=2, values_only=True)]
        assert opp_titles == ["Keeper TPM", "Undecided"]
        # S.No renumbered sequentially on the survivors.
        opp_snos = [r[0] for r in wb["Opportunities"].iter_rows(min_row=2, values_only=True)]
        assert opp_snos == [1, 2]

        rej = list(wb["Rejected opportunities"].iter_rows(min_row=2, values_only=True))
        rej_titles = sorted(r[1] for r in rej)
        assert rej_titles == ["Invalid", "Not Worth"]


def test_reason_lands_in_reason_column_by_name():
    # Even though the rejected sheet header is narrower/wider than Opportunities,
    # the reason must map to the 'Reason' column by name, not by position.
    with tempfile.TemporaryDirectory() as d:
        path = os.path.join(d, "opps.xlsx")
        _make_workbook(path, [_row(1, "Not Worth", "MehCo", worth="N", valid="Y")])
        purge(path)
        wb = load_workbook(path)
        rej = wb["Rejected opportunities"]
        header = [c.value for c in rej[1]]
        reason_idx = [h.lower() for h in header].index("reason")
        first = [c.value for c in rej[2]]
        assert first[reason_idx] == "Worth Applying = N"
        assert "rejected on" in [h.lower() for h in header]


def test_sync_rejected_header_widens_and_dedups():
    out = sync_rejected_header(["S.No", "Title", "Applied"], ["S.No", "Title"])
    assert out[:2] == ["S.No", "Title"]
    assert "Applied" in out and "Rejected on" in out and "Reason" in out
    # No duplicates regardless of case.
    lowered = [h.lower() for h in out]
    assert len(lowered) == len(set(lowered))


def test_build_rejected_row_maps_by_name():
    opp_header = ["S.No", "Title", "Company name"]
    rej_header = ["Title", "Company name", "Rejected on", "Reason"]
    row = build_rejected_row(opp_header, [7, "TPM", "Acme"], rej_header, "2026-01-01", "Valid = N")
    assert row == ["TPM", "Acme", "2026-01-01", "Valid = N"]


if __name__ == "__main__":
    fns = [v for k, v in sorted(globals().items()) if k.startswith("test_") and callable(v)]
    failures = []
    for fn in fns:
        try:
            fn()
        except AssertionError as e:
            failures.append(f"{fn.__name__}: {e}")
    if failures:
        print("FAILED:")
        for f in failures:
            print("  -", f)
        sys.exit(1)
    print(f"all {len(fns)} purge_opportunities tests passed")
