#!/usr/bin/env python3
"""Classify discovered candidates against the roles you've already handled.

This is the folder/tracker dedup layer. The discovery engine (`discover.py`) can
already skip roles listed in a lightweight `seen.json`; this goes further and
matches candidates against your real spreadsheets and on-disk folders, so a role
never resurfaces once you've applied to it OR reviewed and declined it.

Dedup sources, in confidence order:
  1. Application tracker (xlsx) — Title + Link columns give per-role dedup,
     checked in three layers:
       a. Raw URL exact match against the tracker's Link.
       b. ATS-native job ID match extracted from the URL — catches the same role
          surfaced on a different URL surface (e.g. a Greenhouse boards URL vs.
          the company's careers page proxying the same posting).
       c. Company + normalized-title match — the fallback when two URL surfaces
          don't share an identifier (e.g. a careers URL vs. its LinkedIn mirror).
  2. `Rejected opportunities` sheet inside the opportunities xlsx — roles you
     reviewed and marked not worth applying to. Checked in the same three layers,
     because the company name there is hand-typed and drifts from what the
     scrapers emit, so the identifier layers hold when name matching doesn't.
  3. Per-company folders on disk — dedup by JD filename + best-effort body parse.

Verdicts:
  - `exact_match`  — definitively the same role. Drop it. `match_source` names the
    layer that fired (`tracker_url`, `tracker_ats_id`, `tracker_title`,
    `rejected_opps_url`, `rejected_opps_ats_id`, `rejected_opps_title`,
    `folder_title`).
  - `company_match` — you've engaged this company but no exact-role signal
    matched. Surface it flagged "verify this is a different role."
    `match_source` is `tracker_company` or `folder_company`.
  - `fresh` — company not seen. Surface clean.

Input: candidate JSON (a list, or an object with a `candidates` key) on stdin or
via one or more `--candidates <path>`. This matches `discover.py`'s output.
Output: JSON to stdout; `--pretty` formats it.

Run:
  python3 discover.py > candidates.json
  python3 dedup_check.py --candidates candidates.json --pretty
"""
from __future__ import annotations

import argparse
import json
import os
import re
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from normalize import normalize_company, normalize_title  # noqa: E402
from read_jds import collect_jds  # noqa: E402


# ---------------------------------------------------------------------------
# ATS job-ID extraction — the same role can surface on several URLs. Greenhouse
# `gh_jid=<n>` appears both on greenhouse.io boards and on careers pages that
# proxy Greenhouse; Lever/Ashby UUIDs are unique per role. LinkedIn IDs live in
# their own namespace — a LinkedIn ID never matches a Greenhouse ID.
# ---------------------------------------------------------------------------
_ATS_RX_GREENHOUSE_QP = re.compile(r"[?&]gh_jid=(\d+)")
_ATS_RX_GREENHOUSE_PATH = re.compile(r"greenhouse\.io/[^/]+/jobs/(\d+)")
_ATS_RX_LEVER = re.compile(r"lever\.co/[^/]+/([0-9a-f\-]{20,})", re.IGNORECASE)
_ATS_RX_ASHBY = re.compile(r"ashbyhq\.com/[^/]+/([0-9a-f\-]{20,})", re.IGNORECASE)
_ATS_RX_LINKEDIN = re.compile(r"linkedin\.com/jobs/view/(\d+)")


def extract_ats_id(url: str) -> tuple[str | None, str | None]:
    """Return (ats_family, job_id) if the URL carries one, else (None, None)."""
    if not url:
        return (None, None)
    for family, rx in (
        ("greenhouse", _ATS_RX_GREENHOUSE_QP),
        ("greenhouse", _ATS_RX_GREENHOUSE_PATH),
        ("lever", _ATS_RX_LEVER),
        ("ashby", _ATS_RX_ASHBY),
        ("linkedin", _ATS_RX_LINKEDIN),
    ):
        m = rx.search(url)
        if m:
            jid = m.group(1)
            return (family, jid.lower() if family in ("lever", "ashby") else jid)
    return (None, None)


# ---------------------------------------------------------------------------
# Spreadsheet loaders
# ---------------------------------------------------------------------------

def _header_index(header: list[str], *names: str) -> int | None:
    """First column index whose header matches any of `names` (case-insensitive).
    Lets a tracker use either 'Company' or 'Company name', etc."""
    lowered = [h.lower() for h in header]
    for name in names:
        if name.lower() in lowered:
            return lowered.index(name.lower())
    return None


def load_applied_from_tracker(path: str) -> list[dict]:
    """Read the application tracker and return every row (any status — a rejected
    role still shouldn't resurface). Missing Title/Link cells are tolerated."""
    if not os.path.isfile(path):
        return []
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("dedup_check: openpyxl not installed; skipping tracker load", file=sys.stderr)
        return []
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(h).strip() if h is not None else "" for h in rows[0]]
    ci = _header_index(header, "Company name", "Company")
    ti = _header_index(header, "Title")
    li = _header_index(header, "Link")
    si = _header_index(header, "Status")
    if ci is None:
        return []
    out: list[dict] = []
    for r in rows[1:]:
        if not r or ci >= len(r) or not r[ci]:
            continue
        get = lambda i: str(r[i]).strip() if (i is not None and i < len(r) and r[i]) else ""
        title = get(ti)
        out.append({
            "company": str(r[ci]).strip(),
            "title": title,
            "normalized_title": normalize_title(title) if title else "",
            "url": get(li),
            "status": get(si),
        })
    return out


def load_rejected_opps(path: str) -> list[dict]:
    """Read the `Rejected opportunities` sheet in the opportunities xlsx. Missing
    sheet or file → empty list (dedup silently falls through)."""
    if not os.path.isfile(path):
        return []
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("dedup_check: openpyxl not installed; skipping rejected-opps load", file=sys.stderr)
        return []
    wb = load_workbook(path, read_only=True, data_only=True)
    if "Rejected opportunities" not in wb.sheetnames:
        return []
    ws = wb["Rejected opportunities"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(h).strip() if h is not None else "" for h in rows[0]]
    ci = _header_index(header, "Company name", "Company")
    ti = _header_index(header, "Title")
    li = _header_index(header, "Link")
    ri = _header_index(header, "Reason")
    if ci is None or ti is None:
        return []
    out: list[dict] = []
    for r in rows[1:]:
        if not r or ci >= len(r) or not r[ci]:
            continue
        get = lambda i: str(r[i]).strip() if (i is not None and i < len(r) and r[i]) else ""
        title = get(ti)
        if not title:
            continue
        out.append({
            "company": str(r[ci]).strip(),
            "title": title,
            "normalized_title": normalize_title(title),
            "url": get(li),
            "reason": get(ri),
        })
    return out


# ---------------------------------------------------------------------------
# Classification
# ---------------------------------------------------------------------------

def classify(
    candidates: list[dict],
    folder_data: dict,
    applied_rows: list[dict],
    rejected_opps_rows: list[dict] | None = None,
) -> list[dict]:
    rejected_opps_rows = rejected_opps_rows or []

    tracker_by_url = {a["url"]: a for a in applied_rows if a["url"]}
    tracker_by_ats_id: dict[tuple[str, str], dict] = {}
    tracker_by_company_title: dict[tuple[str, str], dict] = {}
    tracker_companies: set[str] = set()
    for a in applied_rows:
        family, jid = extract_ats_id(a["url"])
        if family and jid:
            tracker_by_ats_id[(family, jid)] = a
        ck = normalize_company(a["company"])
        tracker_companies.add(ck)
        if a["normalized_title"]:
            tracker_by_company_title[(ck, a["normalized_title"])] = a

    rejected_by_key: dict[tuple[str, str], dict] = {}
    rejected_by_url: dict[str, dict] = {}
    rejected_by_ats_id: dict[tuple[str, str], dict] = {}
    for r in rejected_opps_rows:
        ck = normalize_company(r["company"])
        if r["normalized_title"]:
            rejected_by_key[(ck, r["normalized_title"])] = r
        if r["url"]:
            rejected_by_url[r["url"]] = r
            family, jid = extract_ats_id(r["url"])
            if family and jid:
                rejected_by_ats_id[(family, jid)] = r

    # Folder dedup keyed on the normalized company name, since folder names drift
    # on corporate suffixes the same way tracker rows do. First writer wins on a
    # rare normalization collision.
    folder_by_company: dict[str, dict] = {}
    for folder_name, fdata in folder_data.items():
        folder_by_company.setdefault(normalize_company(folder_name), fdata)

    out = []
    for c in candidates:
        company = (c.get("company") or "").strip()
        title = (c.get("title") or "").strip()
        url = (c.get("url") or "").strip()
        norm_title = normalize_title(title)
        company_key = normalize_company(company)

        verdict = match_source = matched_file = None
        matched_row: dict | None = None

        family, jid = extract_ats_id(url)
        ats_key = (family, jid) if family and jid else None
        title_key = (company_key, norm_title) if norm_title else None

        # 1–3: tracker layers (url -> ats id -> company+title)
        if url and url in tracker_by_url:
            verdict, match_source = "exact_match", "tracker_url"
            matched_row = tracker_by_url[url]
        elif ats_key and ats_key in tracker_by_ats_id:
            verdict, match_source = "exact_match", "tracker_ats_id"
            matched_row = tracker_by_ats_id[ats_key]
        elif title_key and title_key in tracker_by_company_title:
            verdict, match_source = "exact_match", "tracker_title"
            matched_row = tracker_by_company_title[title_key]
        # 4: rejected-opportunities layers (same three, same priority)
        elif url and url in rejected_by_url:
            verdict, match_source = "exact_match", "rejected_opps_url"
            matched_row = rejected_by_url[url]
        elif ats_key and ats_key in rejected_by_ats_id:
            verdict, match_source = "exact_match", "rejected_opps_ats_id"
            matched_row = rejected_by_ats_id[ats_key]
        elif title_key and title_key in rejected_by_key:
            verdict, match_source = "exact_match", "rejected_opps_title"
            matched_row = rejected_by_key[title_key]

        # 5: folder layers, then company-presence fallback
        folder = folder_by_company.get(company_key)
        if verdict is None:
            if folder is not None and norm_title in folder["titles"]:
                verdict, match_source = "exact_match", "folder_title"
                matched_file = folder["files"][0] if folder["files"] else None
            elif folder is not None:
                verdict, match_source = "company_match", "folder_company"
            elif company_key in tracker_companies:
                verdict, match_source = "company_match", "tracker_company"
            else:
                verdict = "fresh"

        matched_out = None
        if matched_row:
            matched_out = {
                "title": matched_row["title"],
                "url": matched_row["url"],
                "status": matched_row.get("status", ""),
            }
            # Surface WHY a rejected-opps row was declined, so an audit log can
            # show the reason regardless of which layer matched.
            if (match_source or "").startswith("rejected_opps_") and matched_row.get("reason"):
                matched_out["rejected_reason"] = matched_row["reason"]

        out.append({
            "company": company,
            "title": title,
            "url": url,
            "normalized_title": norm_title,
            "verdict": verdict,
            "match_source": match_source,
            "matched_tracker_row": matched_out,
            "folder_titles": folder["titles"] if folder else [],
            "matched_file": matched_file,
            "location": c.get("location"),
            "location_score": c.get("location_score"),
            "age_days": c.get("age_days"),
            "salary": c.get("salary") or c.get("salary_parsed"),
        })
    return out


def main():
    here = os.path.dirname(os.path.abspath(__file__))
    p = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    p.add_argument("--candidates", action="append",
                   help="Candidate JSON file. Repeatable to merge inputs "
                        "(e.g. ATS + LinkedIn-fallback candidates). Defaults to stdin.")
    p.add_argument("--root", default=here,
                   help="Directory containing per-company folders (default: this script's dir).")
    p.add_argument("--tracker", default=os.path.join(here, "Application tracker.xlsx"),
                   help="Path to the application tracker xlsx (default: sibling file).")
    p.add_argument("--opps", default=os.path.join(here, "TPM opportunities.xlsx"),
                   help="Path to the opportunities xlsx; its 'Rejected opportunities' "
                        "sheet is loaded as an extra exact-match source (default: sibling file).")
    p.add_argument("--pretty", action="store_true", help="Pretty-print JSON")
    args = p.parse_args()

    candidates: list[dict] = []
    if args.candidates:
        for path in args.candidates:
            with open(path) as f:
                raw = json.load(f)
            candidates.extend(raw.get("candidates", raw) if isinstance(raw, dict) else raw)
    else:
        raw = json.load(sys.stdin)
        candidates = raw.get("candidates", raw) if isinstance(raw, dict) else raw

    folder_data = collect_jds(os.path.abspath(args.root))
    applied_rows = load_applied_from_tracker(os.path.abspath(args.tracker))
    rejected_opps_rows = load_rejected_opps(os.path.abspath(args.opps))
    classified = classify(candidates, folder_data, applied_rows, rejected_opps_rows)

    summary = {
        "exact_match": sum(1 for c in classified if c["verdict"] == "exact_match"),
        "company_match": sum(1 for c in classified if c["verdict"] == "company_match"),
        "fresh": sum(1 for c in classified if c["verdict"] == "fresh"),
        "total": len(classified),
        "tracker_rows_loaded": len(applied_rows),
        "rejected_opps_rows_loaded": len(rejected_opps_rows),
    }
    print(json.dumps({"summary": summary, "candidates": classified},
                     indent=2 if args.pretty else None, ensure_ascii=False))


if __name__ == "__main__":
    main()
